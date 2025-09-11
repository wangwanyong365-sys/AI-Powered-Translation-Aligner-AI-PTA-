import os
import csv
import json
import time
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog

import openai
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

from app_utils import log_error, save_settings, split_text_into_paragraphs, translate_single_paragraph

RESUME_PE_FILE = "resume_post_edit.json"

class TermEditDialog(tk.Toplevel):

    def __init__(self, parent, title, source_term="", target_term=""):
        super().__init__(parent)
        self.transient(parent)
        self.title(title)
        self.parent = parent
        self.result = None
    
        body = ttk.Frame(self)
        self.initial_focus = self._create_widgets(body, source_term, target_term)
        body.pack(padx=15, pady=15)
    
        self._create_buttons()
    
        self.grab_set()
        if not self.initial_focus:
            self.initial_focus = self
    
        self.protocol("WM_DELETE_WINDOW", self._cancel)
        self.geometry(f"+{parent.winfo_rootx()+50}+{parent.winfo_rooty()+50}")
        self.initial_focus.focus_set()
        self.wait_window(self)
    
    def _create_widgets(self, master, source_term, target_term):
        ttk.Label(master, text="Source Term:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.source_entry = ttk.Entry(master, width=30)
        self.source_entry.grid(row=0, column=1, sticky="ew", padx=5, pady=5)
        self.source_entry.insert(0, source_term)
    
        ttk.Label(master, text="Target Term:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.target_entry = ttk.Entry(master, width=30)
        self.target_entry.grid(row=1, column=1, sticky="ew", padx=5, pady=5)
        self.target_entry.insert(0, target_term)
        
        return self.source_entry
    
    def _create_buttons(self):
        button_frame = ttk.Frame(self)
        
        ok_button = ttk.Button(button_frame, text="OK", command=self._ok)
        ok_button.pack(side="left", padx=5, pady=5)
        
        cancel_button = ttk.Button(button_frame, text="Cancel", command=self._cancel)
        cancel_button.pack(side="left", padx=5, pady=5)
        
        button_frame.pack()
        
        self.bind("<Return>", self._ok)
        self.bind("<Escape>", self._cancel)
    
    def _ok(self, event=None):
        source = self.source_entry.get().strip()
        target = self.target_entry.get().strip()
        if not source or not target:
            messagebox.showwarning("Input Error", "Source term and target term cannot be empty.", parent=self)
            return
    
        self.result = (source, target)
        self.destroy()
    
    def _cancel(self, event=None):
        self.destroy()

class TermAnnotatorApp:

    def __init__(self, root):
        self.root = root
        self.current_terms = {}
        self.source_file_path = tk.StringVar()
    
        self._setup_window()
        self._setup_styles()
        self._create_widgets()
        self._load_terminologies()
    
    def _setup_window(self):
        self.root.title("Source Text Term Annotator")
        self.root.geometry("1000x700")
        self.root.minsize(800, 600)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        self.root.protocol("WM_DELETE_WINDOW", self._on_closing)
    
    def _setup_styles(self):
        self.style = ttk.Style(self.root)
        self.style.theme_use("clam")
        
        default_font = ("Segoe UI", 10)
        self.style.configure("TLabel", font=default_font)
        self.style.configure("TButton", font=default_font, padding=5)
        self.style.configure("TEntry", font=default_font)
        self.style.configure("TCombobox", font=default_font)
        
        self.style.configure("TLabelframe.Label", font=("Segoe UI", 12, "bold"))
        
        self.style.configure("Accent.TButton", font=("Segoe UI", 10, "bold"), padding=8)
        self.style.map("Accent.TButton",
                       background=[("active", "#0078D7")],
                       foreground=[("active", "white")])
    
    def _create_widgets(self):
        main_frame = ttk.Frame(self.root, padding=15)
        main_frame.grid(row=0, column=0, sticky="nsew")
        main_frame.rowconfigure(1, weight=1)
        main_frame.columnconfigure(0, weight=1)
    
        top_controls_frame = self._create_top_controls(main_frame)
        top_controls_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
    
        text_area_frame = self._create_text_areas(main_frame)
        text_area_frame.grid(row=1, column=0, sticky="nsew")
    
        bottom_frame = ttk.Frame(main_frame)
        bottom_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(10, 0))
        bottom_frame.columnconfigure(0, weight=1)
    
        self.annotate_button = ttk.Button(
            bottom_frame, text="Start Annotation", command=self._start_annotation, style="Accent.TButton"
        )
        self.annotate_button.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E))
    
        self.status_label = ttk.Label(self.root, text="Ready", relief=tk.SUNKEN, anchor=tk.W, padding=5)
        self.status_label.grid(row=1, column=0, sticky="ew")
        self._update_status("Ready", "gray")
    
    def _create_top_controls(self, parent):
        frame = ttk.LabelFrame(parent, text="Settings")
        frame.columnconfigure(1, weight=1)
    
        ttk.Label(frame, text="Select Terminology:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.term_db_combo = ttk.Combobox(frame, state="readonly")
        self.term_db_combo.grid(row=0, column=1, padx=5, pady=5, sticky=(tk.W, tk.E))
        self.term_db_combo.bind("<<ComboboxSelected>>", self._on_term_db_selected)
    
        ttk.Label(frame, text="Current Terms:").grid(row=1, column=0, padx=5, pady=5, sticky=(tk.W, tk.N))
        term_list_frame = ttk.Frame(frame)
        term_list_frame.grid(row=1, column=1, rowspan=2, padx=5, pady=5, sticky="nsew")
        term_list_frame.rowconfigure(0, weight=1)
        term_list_frame.columnconfigure(0, weight=1)
        
        self.term_listbox = tk.Listbox(term_list_frame, font=("Segoe UI", 10), height=5)
        self.term_listbox.grid(row=0, column=0, sticky="nsew")
        
        scrollbar = ttk.Scrollbar(term_list_frame, orient=tk.VERTICAL, command=self.term_listbox.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.term_listbox.config(yscrollcommand=scrollbar.set)
        
        frame.rowconfigure(1, weight=1)
    
        term_actions_frame = ttk.Frame(frame)
        term_actions_frame.grid(row=2, column=0, padx=5, pady=5, sticky=tk.W)
        self.add_term_button = ttk.Button(term_actions_frame, text="Add", command=self._add_term, state=tk.DISABLED)
        self.add_term_button.pack(side=tk.LEFT, padx=(0, 5))
        self.modify_term_button = ttk.Button(term_actions_frame, text="Modify", command=self._modify_term, state=tk.DISABLED)
        self.modify_term_button.pack(side=tk.LEFT, padx=(0, 5))
        self.delete_term_button = ttk.Button(term_actions_frame, text="Delete", command=self._delete_term, state=tk.DISABLED)
        self.delete_term_button.pack(side=tk.LEFT)
    
        ttk.Label(frame, text="Select Source File:").grid(row=3, column=0, padx=5, pady=5, sticky=tk.W)
        entry = ttk.Entry(frame, textvariable=self.source_file_path, state="readonly")
        entry.grid(row=3, column=1, padx=5, pady=5, sticky=(tk.W, tk.E))
        browse_button = ttk.Button(frame, text="Browse...", command=self._browse_source_file)
        browse_button.grid(row=3, column=2, padx=5, pady=5, sticky=tk.E)
    
        return frame
    
    def _create_text_areas(self, parent):
        frame = ttk.LabelFrame(parent, text="Text Content")
        frame.rowconfigure(1, weight=1)
        frame.columnconfigure(0, weight=1)
        frame.columnconfigure(1, weight=1)
    
        ttk.Label(frame, text="Source Text").grid(row=0, column=0, padx=5, pady=5)
        source_text_frame = ttk.Frame(frame)
        source_text_frame.grid(row=1, column=0, sticky="nsew", padx=(0, 5))
        source_text_frame.rowconfigure(0, weight=1)
        source_text_frame.columnconfigure(0, weight=1)
        
        self.source_text = tk.Text(source_text_frame, wrap=tk.WORD, font=("Segoe UI", 10), undo=True)
        self.source_text.grid(row=0, column=0, sticky="nsew")
        source_scrollbar = ttk.Scrollbar(source_text_frame, orient=tk.VERTICAL, command=self.source_text.yview)
        source_scrollbar.grid(row=0, column=1, sticky="ns")
        self.source_text.config(yscrollcommand=source_scrollbar.set)
    
        ttk.Label(frame, text="Annotated Text").grid(row=0, column=1, padx=5, pady=5)
        annotated_text_frame = ttk.Frame(frame)
        annotated_text_frame.grid(row=1, column=1, sticky="nsew", padx=(5, 0))
        annotated_text_frame.rowconfigure(0, weight=1)
        annotated_text_frame.columnconfigure(0, weight=1)
    
        self.annotated_text = tk.Text(annotated_text_frame, wrap=tk.WORD, font=("Segoe UI", 10), undo=True, state=tk.DISABLED)
        self.annotated_text.grid(row=0, column=0, sticky="nsew")
        annotated_scrollbar = ttk.Scrollbar(annotated_text_frame, orient=tk.VERTICAL, command=self.annotated_text.yview)
        annotated_scrollbar.grid(row=0, column=1, sticky="ns")
        self.annotated_text.config(yscrollcommand=annotated_scrollbar.set)
        
        export_button = ttk.Button(frame, text="Export as .txt", command=self._export_annotated_text)
        export_button.grid(row=2, column=1, sticky=tk.E, padx=5, pady=(10, 0))
    
        return frame
    
    def _update_status(self, text, color="black"):
        self.status_label.config(text=text, foreground=color)
        self.root.update_idletasks()
    
    def _load_terminologies(self):
        term_dir = "terminology"
        if not os.path.exists(term_dir):
            os.makedirs(term_dir)
            messagebox.showinfo("Info", f"The 'terminology' folder has been created.\nPlease put your CSV terminology files in it.", parent=self.root)
            self._update_status("Terminology folder created. Please add files.", "blue")
            return
    
        try:
            csv_files = [f for f in os.listdir(term_dir) if f.endswith('.csv')]
            self.term_db_combo['values'] = csv_files
            if csv_files:
                self.term_db_combo.current(0)
                self._on_term_db_selected(None)
                self._update_status(f"Loaded {len(csv_files)} terminologies.", "green")
            else:
                self._update_status("No .csv files found in the 'terminology' folder.", "orange")
        except Exception as e:
            messagebox.showerror("Error", f"Error loading terminologies: {e}", parent=self.root)
            self._update_status(f"Failed to load terminologies: {e}", "red")
    
    def _on_term_db_selected(self, event):
        filename = self.term_db_combo.get()
        if not filename: return
        
        filepath = os.path.join("terminology", filename)
        self.current_terms = {}
        self.term_listbox.delete(0, tk.END)
    
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for i, row in enumerate(reader):
                    if len(row) >= 2 and row[0] and row[1]:
                        source_term, target_term = row[0].strip(), row[1].strip()
                        self.current_terms[source_term] = target_term
                    else:
                        print(f"Warning: Skipping invalid row {i+1} in file '{filename}': {row}")
            self._update_term_listbox()
            self._update_status(f"Successfully loaded '{filename}' with {len(self.current_terms)} terms.", "green")
            for btn in [self.add_term_button, self.modify_term_button, self.delete_term_button]:
                btn.config(state=tk.NORMAL)
        except FileNotFoundError:
            messagebox.showerror("Error", f"File '{filename}' not found.", parent=self.root)
            self._update_status(f"File not found: {filename}", "red")
        except Exception as e:
            messagebox.showerror("Error", f"Error reading file '{filename}': {e}", parent=self.root)
            self._update_status(f"Failed to read file '{filename}'", "red")
    
    def _update_term_listbox(self):
        self.term_listbox.delete(0, tk.END)
        for source, target in sorted(self.current_terms.items()):
            self.term_listbox.insert(tk.END, f"{source} → {target}")
    
    def _save_current_terms(self):
        filename = self.term_db_combo.get()
        if not filename:
            self._update_status("Error: No terminology file selected for saving.", "red")
            return False
        
        filepath = os.path.join("terminology", filename)
        try:
            with open(filepath, 'w', encoding='utf-8', newline='') as f:
                writer = csv.writer(f)
                for source, target in sorted(self.current_terms.items()):
                    writer.writerow([source, target])
            self._update_status(f"Terminology '{filename}' saved automatically.", "green")
            return True
        except Exception as e:
            messagebox.showerror("Save Failed", f"Could not save terminology '{filename}':\n{e}", parent=self.root)
            self._update_status(f"Failed to save terminology: {e}", "red")
            return False
    
    def _add_term(self):
        dialog = TermEditDialog(self.root, "Add New Term")
        if dialog.result:
            source, target = dialog.result
            if source in self.current_terms:
                if not messagebox.askyesno("Term Exists", f"Source term '{source}' already exists. Do you want to overwrite it?", parent=self.root):
                    return
            
            self.current_terms[source] = target
            if self._save_current_terms():
                self._update_term_listbox()
                self._update_status(f"Added term: {source} → {target}", "green")
    
    def _modify_term(self):
        selected_indices = self.term_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("Invalid Operation", "Please select a term to modify from the list first.", parent=self.root)
            return
    
        selected_item = self.term_listbox.get(selected_indices[0])
        try:
            old_source, old_target = [s.strip() for s in selected_item.split('→')]
        except ValueError:
            messagebox.showerror("Format Error", "Could not parse the selected term line.", parent=self.root)
            return
    
        dialog = TermEditDialog(self.root, "Modify Term", old_source, old_target)
        if dialog.result:
            new_source, new_target = dialog.result
            del self.current_terms[old_source]
            self.current_terms[new_source] = new_target
            
            if self._save_current_terms():
                self._update_term_listbox()
                self._update_status(f"Modified term: {new_source} → {new_target}", "green")
            else:
                del self.current_terms[new_source]
                self.current_terms[old_source] = old_target
    
    def _delete_term(self):
        selected_indices = self.term_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("Invalid Operation", "Please select a term to delete from the list first.", parent=self.root)
            return
            
        selected_item = self.term_listbox.get(selected_indices[0])
        if messagebox.askyesno("Confirm Deletion", f"Are you sure you want to delete the following term?\n\n{selected_item}", parent=self.root):
            try:
                source_to_delete = selected_item.split('→')[0].strip()
                if source_to_delete in self.current_terms:
                    del self.current_terms[source_to_delete]
                    if self._save_current_terms():
                        self._update_term_listbox()
                        self._update_status(f"Deleted term: {source_to_delete}", "green")
            except Exception as e:
                messagebox.showerror("Deletion Failed", f"An error occurred during deletion: {e}", parent=self.root)
    
    def _browse_source_file(self):
        filepath = filedialog.askopenfilename(
            title="Please select the source text file",
            filetypes=(("Text files", "*.txt"), ("All files", "*.*")),
            parent=self.root  
        )
        if filepath:
            self.source_file_path.set(filepath)
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    self.source_text.delete("1.0", tk.END)
                    self.source_text.insert("1.0", f.read())
                self._update_status(f"Loaded source file: {os.path.basename(filepath)}", "blue")
            except Exception as e:
                messagebox.showerror("File Read Error", f"Could not read file: {e}", parent=self.root)
                self._update_status(f"File read failed: {e}", "red")
    
    def _export_annotated_text(self):
        content = self.annotated_text.get("1.0", tk.END).strip()
        if not content:
            messagebox.showwarning("Invalid Operation", "No annotated text to export.", parent=self.root)
            return
        
        source_path = self.source_file_path.get()
        default_filename = "annotated_text.txt"
        if source_path:
            base, ext = os.path.splitext(os.path.basename(source_path))
            default_filename = f"{base}_annotated{ext}"
    
        filepath = filedialog.asksaveasfilename(
            title="Export Annotated Text",
            defaultextension=".txt",
            filetypes=(("Text files", "*.txt"), ("All files", "*.*")),
            initialfile=default_filename,
            parent=self.root
        )
    
        if filepath:
            try:
                with open(filepath, 'w', encoding='utf-8') as f:
                    f.write(content)
                self._update_status(f"File successfully exported to: {os.path.basename(filepath)}", "green")
            except Exception as e:
                messagebox.showerror("Export Failed", f"Could not write to file: {e}", parent=self.root)
                self._update_status(f"File export failed: {e}", "red")
    
    def _perform_annotation(self, source_text, term_dict):
        sorted_terms = sorted(term_dict.keys(), key=len, reverse=True)
        annotated_text = source_text
        for source_term in sorted_terms:
            target_term = term_dict[source_term]
            annotation_tag = f"{source_term}{{{target_term}}}"
            annotated_text = annotated_text.replace(source_term, annotation_tag)
        return annotated_text
    
    def _start_annotation(self):
        source_text = self.source_text.get("1.0", tk.END).strip()
        if not self.current_terms:
            messagebox.showwarning("Invalid Operation", "Please select and load a valid terminology first.", parent=self.root)
            return
        if not source_text:
            messagebox.showwarning("Invalid Operation", "Source text content cannot be empty.", parent=self.root)
            return
    
        self.annotate_button.config(state=tk.DISABLED)
        self._update_status("Annotating, please wait...", "orange")
        
        try:
            result = self._perform_annotation(source_text, self.current_terms)
            self.annotated_text.config(state=tk.NORMAL)
            self.annotated_text.delete("1.0", tk.END)
            self.annotated_text.insert("1.0", result)
            self.annotated_text.config(state=tk.DISABLED)
            self._update_status("Annotation complete!", "green")
        except Exception as e:
            self._update_status(f"An error occurred during annotation: {e}", "red")
            messagebox.showerror("Annotation Failed", f"An unknown error occurred: {e}", parent=self.root)
        finally:
            self.annotate_button.config(state=tk.NORMAL)
    
    def _on_closing(self):
        self.root.destroy()

class PostEditingWindow(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.title("Post-editing Tool")
        self.geometry("700x600")
        self.minsize(600, 500)
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        self.source_file_path = tk.StringVar()
        self.stop_requested = threading.Event()
        self.is_processing = False
        self.resume_data = None
        self.timer_id = None
    
        self._setup_style()
        self._setup_ui()
        self._post_ui_setup()
        self.protocol("WM_DELETE_WINDOW", self._on_closing)
        self.transient(parent)
        self.grab_set()
        self.after(100, self._check_for_resume_task)
    
    def _on_closing(self):
        if self.is_processing:
            if not messagebox.askyesno("Confirm Exit", "A post-editing task is in progress. Exiting now will stop it. Are you sure?", parent=self):
                return
            self.stop_requested.set()
        self.destroy()

    def _setup_style(self):
        self.style = ttk.Style(self)
        self.style.theme_use("clam")
        
        default_font = ("Segoe UI", 10)
        self.style.configure("TLabel", font=default_font)
        self.style.configure("TButton", font=default_font, padding=5)
        self.style.configure("TEntry", font=default_font)
        
        self.style.configure("TLabelframe.Label", font=("Segoe UI", 12, "bold"))
        
        self.style.configure("Accent.TButton", font=("Segoe UI", 10, "bold"), padding=8)
        self.style.map("Accent.TButton",
                       background=[("active", "#0078D7")],
                       foreground=[("active", "white")])
    
    def _setup_ui(self):
        main_frame = ttk.Frame(self, padding="15")
        main_frame.grid(row=0, column=0, sticky="nsew")
        main_frame.rowconfigure(1, weight=1)
        main_frame.columnconfigure(0, weight=1)
    
        file_frame = ttk.LabelFrame(main_frame, text="Source File", padding="10")
        file_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        file_frame.columnconfigure(0, weight=1)
        
        entry = ttk.Entry(file_frame, textvariable=self.source_file_path, state="readonly")
        entry.grid(row=0, column=0, sticky="ew", padx=(0, 5))
        browse_button = ttk.Button(file_frame, text="Browse...", command=self._browse_file)
        browse_button.grid(row=0, column=1)
    
        prompt_frame = ttk.LabelFrame(main_frame, text="Post-editing Prompt", padding="10")
        prompt_frame.grid(row=1, column=0, sticky="nsew", pady=5)
        prompt_frame.columnconfigure(0, weight=1)
        prompt_frame.rowconfigure(1, weight=1)
        
        prompt_selection_frame = ttk.Frame(prompt_frame)
        prompt_selection_frame.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 5))
        prompt_selection_frame.columnconfigure(1, weight=1)
        
        ttk.Label(prompt_selection_frame, text="Select Prompt:").grid(row=0, column=0, padx=(0, 5))
        self.prompt_var = tk.StringVar()
        self.prompt_combo = ttk.Combobox(prompt_selection_frame, textvariable=self.prompt_var, state="readonly")
        self.prompt_combo.grid(row=0, column=1, sticky="ew")
        self.prompt_combo.bind("<<ComboboxSelected>>", self._on_prompt_select)
        
        btn_frame = ttk.Frame(prompt_selection_frame)
        btn_frame.grid(row=0, column=2, padx=(10, 0))
        ttk.Button(btn_frame, text="Add", command=self._add_prompt, width=6).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="Save", command=self._save_current_prompt, width=6).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="Delete", command=self._delete_prompt, width=6).pack(side=tk.LEFT, padx=2)
        
        self.prompt_text = tk.Text(prompt_frame, wrap=tk.WORD, height=8, font=("Segoe UI", 10))
        self.prompt_text.grid(row=1, column=0, sticky="nsew")
        prompt_scrollbar = ttk.Scrollbar(prompt_frame, orient=tk.VERTICAL, command=self.prompt_text.yview)
        prompt_scrollbar.grid(row=1, column=1, sticky="ns")
        self.prompt_text.config(yscrollcommand=prompt_scrollbar.set)
    
        self.process_button = ttk.Button(main_frame, text="Start Post-editing", command=self._start_post_editing, style="Accent.TButton")
        self.process_button.grid(row=2, column=0, pady=10, sticky="ew")
    
        status_bar = ttk.Frame(self)
        status_bar.grid(row=1, column=0, sticky="ew")
        self.status_label = ttk.Label(status_bar, text="Ready", relief=tk.SUNKEN, anchor=tk.W, padding=5)
        self.status_label.pack(side=tk.LEFT, expand=True, fill=tk.X)
        self.timer_label = ttk.Label(status_bar, text="", relief=tk.SUNKEN, anchor=tk.E, padding=5)
        self.timer_label.pack(side=tk.RIGHT)
        self._update_status("Ready", "gray")
    
    def _post_ui_setup(self):
        self._update_prompt_combo()
        if self.parent.settings['post_editing_prompts']:
            first_prompt_name = list(self.parent.settings['post_editing_prompts'].keys())[0]
            self.prompt_var.set(first_prompt_name)
            self._on_prompt_select()
    
    def _browse_file(self):
        filepath = filedialog.askopenfilename(
            title="Select the Excel file to post-edit",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")),
            parent=self
        )
        if filepath:
            self.source_file_path.set(filepath)
            self._update_status(f"Selected file: {os.path.basename(filepath)}", "blue")
    
    def _update_status(self, text, color):
        self.status_label.config(text=text, foreground=color)
        self.update_idletasks()
    
    def _update_timer(self, start_time):
        elapsed = time.time() - start_time
        mins, secs = divmod(elapsed, 60)
        self.timer_label.config(text=f"{int(mins):02d}:{secs:04.1f}")
        self.timer_id = self.after(100, self._update_timer, start_time)
    
    def _cancel_timer(self):
        if self.timer_id:
            self.after_cancel(self.timer_id)
            self.timer_id = None
        self.timer_label.config(text="")
    
    def _update_prompt_combo(self):
        self.prompt_combo['values'] = list(self.parent.settings['post_editing_prompts'].keys())
    
    def _on_prompt_select(self, event=None):
        name = self.prompt_var.get()
        if name in self.parent.settings['post_editing_prompts']:
            self.prompt_text.delete("1.0", tk.END)
            self.prompt_text.insert("1.0", self.parent.settings['post_editing_prompts'][name])
    
    def _add_prompt(self):
        name = simpledialog.askstring("New Prompt", "Enter name for the new Post-editing Prompt:", parent=self)
        if name and name.strip():
            name = name.strip()
            if name in self.parent.settings['post_editing_prompts']:
                messagebox.showerror("Error", "This name already exists!", parent=self)
                return
            self.parent.settings['post_editing_prompts'][name] = "Enter instructions, using {source} and {target} placeholders."
            save_settings(self.parent.settings)
            self._update_prompt_combo()
            self.prompt_var.set(name)
            self._on_prompt_select()
    
    def _save_current_prompt(self):
        name = self.prompt_var.get()
        if not name: return messagebox.showerror("Error", "No prompt selected to save.", parent=self)
        content = self.prompt_text.get("1.0", tk.END).strip()
        if "{source}" not in content or "{target}" not in content:
            if not messagebox.askyesno("Warning", "The placeholders {source} and {target} were not found. This may cause errors.\nDo you still want to save?", parent=self):
                return
        self.parent.settings['post_editing_prompts'][name] = content
        save_settings(self.parent.settings)
        messagebox.showinfo("Success", f"Prompt '{name}' has been saved.", parent=self)
    
    def _delete_prompt(self):
        name = self.prompt_var.get()
        if not name: return messagebox.showerror("Error", "Please select a prompt to delete.", parent=self)
        if len(self.parent.settings['post_editing_prompts']) <= 1:
            return messagebox.showwarning("Warning", "You cannot delete the last prompt.", parent=self)
        if messagebox.askyesno("Confirm Deletion", f"Are you sure you want to delete prompt '{name}'?", parent=self):
            del self.parent.settings['post_editing_prompts'][name]
            save_settings(self.parent.settings)
            self._update_prompt_combo()
            first_name = list(self.parent.settings['post_editing_prompts'].keys())[0]
            self.prompt_var.set(first_name)
            self._on_prompt_select()
    
    def _check_for_resume_task(self):
        if os.path.exists(RESUME_PE_FILE):
            try:
                with open(RESUME_PE_FILE, 'r', encoding='utf-8') as f: data = json.load(f)
                file_name = os.path.basename(data.get('current_file', ''))
                row_index = data.get('last_row_index', -1)
                if messagebox.askyesno("Unfinished Task", f"Found an unfinished post-editing task for '{file_name}' (stopped at row {row_index + 1}).\nResume?", parent=self):
                    self._load_resume_state(data)
                else:
                    os.remove(RESUME_PE_FILE)
            except Exception as e:
                log_error(f"Failed to read post-edit resume file: {e}")
                if os.path.exists(RESUME_PE_FILE): os.remove(RESUME_PE_FILE)
    
    def _load_resume_state(self, data):
        self.resume_data = data
        self.source_file_path.set(data.get('current_file', ''))
        self._update_status("Ready to resume. Click 'Start Post-editing'.", "blue")
    
    def _save_resume_state(self, current_file, last_index, edited_paras):
        state = {'current_file': current_file, 'last_row_index': last_index, 'edited_paragraphs': edited_paras}
        try:
            with open(RESUME_PE_FILE, 'w', encoding='utf-8') as f: json.dump(state, f, indent=4)
        except Exception as e:
            log_error(f"Failed to save post-edit resume state: {e}")
    
    def _start_post_editing(self):
        if not self.source_file_path.get(): return messagebox.showerror("Error", "Please select an Excel file.", parent=self)
        prompt = self.prompt_text.get("1.0", tk.END).strip()
        if not prompt: return messagebox.showerror("Error", "Prompt cannot be empty.", parent=self)
        if "{source}" not in prompt or "{target}" not in prompt:
            return messagebox.showerror("Error", "Prompt must contain {source} and {target} placeholders.", parent=self)
    
        self.is_processing = True
        self.stop_requested.clear()
        self.process_button.config(text="Stop Processing", command=self._stop_post_editing)
        self._update_status("Processing...", "orange")
    
        threading.Thread(target=self._post_editing_task, args=(self.resume_data,), daemon=True).start()
        self.resume_data = None
    
    def _stop_post_editing(self):
        self.stop_requested.set()
        self._cancel_timer()
        self._update_status("Stopping...", "orange")
    
    def _post_editing_task(self, resume_data=None):
        try:
            model_name = self.parent.model_name_var.get().strip()
            max_tokens = self.parent.settings.get('max_tokens', 8000)
            retry_attempts = self.parent.settings.get('retry_attempts', 3)
            paragraph_timeout = self.parent.settings.get('paragraph_timeout', 300)
            prompt_template = self.prompt_text.get("1.0", tk.END).strip()
            
            client = self.parent._create_client()
            
            file_path = self.source_file_path.get()
            
            self.after(0, self._update_status, f"Reading: {os.path.basename(file_path)}", "orange")
            df = pd.read_excel(file_path)
            
            if 'Source' not in df.columns or 'Translation' not in df.columns:
                raise ValueError("Excel file must contain 'Source' and 'Translation' columns.")
    
            edited_paragraphs = []
            total_rows = len(df)
            start_row = 0
            if resume_data and file_path == resume_data.get('current_file'):
                start_row = resume_data.get('last_row_index', -1) + 1
                edited_paragraphs = resume_data.get('edited_paragraphs', [])
    
            for i, row in df.iloc[start_row:].iterrows():
                if self.stop_requested.is_set():
                    self._save_resume_state(file_path, i - 1, edited_paragraphs)
                    self.after(0, self._update_status, f"Stopped. Progress for '{os.path.basename(file_path)}' saved.", "blue")
                    return
    
                self.after(0, self._update_status, f"Editing row {i + 1}/{total_rows}", "orange")
                start_time = time.time()
                self.after(0, self._update_timer, start_time)
    
                source_text, target_text = str(row['Source']), str(row['Translation'])
                full_prompt = prompt_template.format(source=source_text, target=target_text)
                
                edited_para = translate_single_paragraph(client, model_name, full_prompt, max_tokens, retry_attempts, paragraph_timeout)
                
                self.after(0, self._cancel_timer)
                edited_paragraphs.append(edited_para)
            
            self.after(0, self._update_status, "Saving output files...", "orange")
    
            df['Post-edited'] = pd.Series(edited_paragraphs, index=df.index[start_row:start_row + len(edited_paragraphs)])
            output_df = df
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            output_dir = os.path.dirname(file_path)
    
            excel_path = os.path.join(output_dir, f"{base_name}_postedited.xlsx")
            output_df.to_excel(excel_path, index=False, engine='openpyxl')
            
            red_bold_font = Font(color="FF0000", bold=True)
            wb = load_workbook(excel_path)
            ws = wb.active
            post_edited_col_idx = -1
            for col_idx, cell in enumerate(ws[1]):
                if cell.value == 'Post-edited':
                    post_edited_col_idx = col_idx
                    break
            
            if post_edited_col_idx != -1:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    cell = row[post_edited_col_idx]
                    if not isinstance(cell.value, str): continue
                    if cell.value == "[ERROR_CONTENT_FILTER]":
                        cell.value = "Rejected by API (content policy)"
                        cell.font = red_bold_font
                    elif cell.value == "[ERROR_NETWORK]":
                        cell.value = "Network Issue"
                        cell.font = red_bold_font
                    elif cell.value.startswith("[ERROR_OTHER:"):
                        cell.value = f"Failed: {cell.value[13:-3]}"
                        cell.font = red_bold_font
            wb.save(excel_path)
            
            full_edited_text = "\n\n".join(output_df['Post-edited'].astype(str).tolist())
            txt_path = os.path.join(output_dir, f"{base_name}_postedited.txt")
            with open(txt_path, 'w', encoding='utf-8') as f: f.write(full_edited_text)
            
            if os.path.exists(RESUME_PE_FILE): os.remove(RESUME_PE_FILE)
            self.after(0, self._update_status, "Post-editing complete! Files saved.", "green")
    
        except Exception as e:
            error_message = f"Processing failed: {e}"
            log_error(f"Post-editing task failed. Error: {e}")
            self.after(0, self._update_status, error_message, "red")
            self.after(0, messagebox.showerror, "An Error Occurred", f"{e}\n\nDetails logged to error_log.txt", parent=self)
        
        finally:
            self.is_processing = False
            self.after(0, lambda: self.process_button.config(text="Start Post-editing", command=self._start_post_editing))
            self.after(0, self._cancel_timer)